"""
VOLTGRID — Energy Monitoring Dashboard (Flask)
------------------------------------------------
Run:
    pip install -r requirements.txt
    python app.py
Then open http://<your-ip>:5000  (accessible across your network).

Roles:
    admin           -> manage users, add/edit data, view everything
    user + can_edit -> add/edit data + view
    user (view-only)-> view only

Default admin (created on first run):  username: admin   password: admin123
*** Change this password after first login. ***

Data source:  data/All_Department_Monthly_Kwh.xlsx
The dashboard reads this file live — any change you make to it (directly in
Excel on the server, or via the in-app "Add / Edit Data" form) shows up
automatically (the UI polls for file changes).
"""
import os, io, threading, json, re
from functools import wraps
from datetime import datetime
from flask import (Flask, request, session, redirect, url_for, render_template,
                   jsonify, send_file, abort)
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
XLSX_PATH = os.path.join(DATA_DIR, "All_Department_Monthly_Kwh.xlsx")
USERS_XLSX = os.path.join(DATA_DIR, "users_config.xlsx")   # user-management data lives here
COLORDER_PATH = os.path.join(DATA_DIR, "column_order.json")  # admin's custom column order per sheet
_cfg_lock = threading.Lock()

app = Flask(__name__)
app.secret_key = os.environ.get("VOLTGRID_SECRET", "change-this-secret-key-in-production")

# --------------------------------------------------------------------------
# User store — persisted to an Excel workbook (data/users_config.xlsx)
# Columns: id | username | pw_hash | role | can_edit | created
# Passwords are stored only as secure hashes (never plain text).
# --------------------------------------------------------------------------
USER_COLS = ["id", "username", "pw_hash", "role", "can_edit", "created"]
_users_lock = threading.Lock()

def _read_users():
    if not os.path.exists(USERS_XLSX):
        return []
    wb = load_workbook(USERS_XLSX, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None for v in r):
            continue
        d = dict(zip(headers, r))
        if not d.get("username"):
            continue
        d["id"] = int(d["id"]) if d.get("id") not in (None, "") else 0
        d["can_edit"] = int(d["can_edit"]) if d.get("can_edit") not in (None, "") else 0
        d["role"] = d.get("role") or "user"
        d["pw_hash"] = d.get("pw_hash") or ""
        d["created"] = str(d.get("created") or "")
        rows.append(d)
    return rows

def _write_users(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(USER_COLS)
    for d in rows:
        ws.append([d.get(c) for c in USER_COLS])
    wb.save(USERS_XLSX)

def _next_id(rows):
    return max([r["id"] for r in rows], default=0) + 1

def init_users():
    os.makedirs(DATA_DIR, exist_ok=True)
    rows = _read_users()
    if not any(r["role"] == "admin" for r in rows):
        rows.append({"id": _next_id(rows), "username": "admin",
                     "pw_hash": generate_password_hash("admin123"),
                     "role": "admin", "can_edit": 1,
                     "created": datetime.now().isoformat(timespec="seconds")})
        _write_users(rows)

def user_by_id(uid):
    return next((r for r in _read_users() if r["id"] == int(uid)), None)

def user_by_name(name):
    return next((r for r in _read_users() if str(r["username"]).lower() == str(name).lower()), None)

# --------------------------------------------------------------------------
# Auth helpers
# --------------------------------------------------------------------------
def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    return user_by_id(uid)

def _wants_json():
    return request.path.startswith("/api/")

def login_required(f):
    @wraps(f)
    def w(*a, **k):
        if not current_user():
            if _wants_json():
                return jsonify(ok=False, error="Your session expired. Please refresh and log in again."), 401
            return redirect(url_for("login"))
        return f(*a, **k)
    return w

def admin_required(f):
    @wraps(f)
    def w(*a, **k):
        u = current_user()
        if not u:
            if _wants_json():
                return jsonify(ok=False, error="Your session expired. Please refresh and log in again."), 401
            return redirect(url_for("login"))
        if u["role"] != "admin":
            if _wants_json():
                return jsonify(ok=False, error="Admin access is required for this action."), 403
            return abort(403)
        return f(*a, **k)
    return w

def can_edit(u):
    return bool(u) and (u["role"] == "admin" or u["can_edit"])

# --------------------------------------------------------------------------
# Excel read / write  (read live so external edits show automatically)
# --------------------------------------------------------------------------
def read_workbook():
    xl = pd.ExcelFile(XLSX_PATH)
    out = {}
    for s in xl.sheet_names:
        dfm = pd.read_excel(xl, sheet_name=s)
        recs = []
        for _, row in dfm.iterrows():
            m = row["Month"]
            rec = {"Month": (str(m)[:7] if not pd.isna(m) else None)}
            for c in dfm.columns:
                if c == "Month":
                    continue
                v = row[c]
                rec[c] = None if pd.isna(v) else round(float(v), 1)
            recs.append(rec)
        out[s] = recs
    return out

def file_version():
    try:
        return int(os.path.getmtime(XLSX_PATH))
    except OSError:
        return 0

def load_colorder():
    try:
        with open(COLORDER_PATH) as f:
            return json.load(f)
    except Exception:
        return {}

def save_colorder(d):
    with open(COLORDER_PATH, "w") as f:
        json.dump(d, f, indent=2)

def write_month_row(sheet, month, values):
    """Update (or create) the row matching YYYY-MM in the given sheet."""
    wb = load_workbook(XLSX_PATH)
    if sheet not in wb.sheetnames:
        raise ValueError("Unknown sheet")
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    if "Month" not in headers:
        raise ValueError("No Month column")
    mcol = headers.index("Month") + 1
    target = None
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=mcol).value
        if cell is None:
            continue
        ym = str(cell)[:7] if not isinstance(cell, datetime) else cell.strftime("%Y-%m")
        if ym == month:
            target = r
            break
    if target is None:  # append a new month row
        target = ws.max_row + 1
        y, mo = month.split("-")
        ws.cell(row=target, column=mcol).value = datetime(int(y), int(mo), 1)
    for col_name, val in values.items():
        if col_name in headers and col_name != "Month":
            c = headers.index(col_name) + 1
            ws.cell(row=target, column=c).value = (None if val in ("", None) else float(val))
    wb.save(XLSX_PATH)

# --------------------------------------------------------------------------
# Routes — pages
# --------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        row = user_by_name(u)
        if row and check_password_hash(row["pw_hash"], p):
            session["uid"] = row["id"]
            return redirect(url_for("index"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def index():
    u = current_user()
    return render_template("dashboard.html", user=u, can_edit=can_edit(u))

@app.route("/users")
@admin_required
def users_page():
    return render_template("users.html", user=current_user())

# --------------------------------------------------------------------------
# Routes — API
# --------------------------------------------------------------------------
@app.route("/api/me")
@login_required
def api_me():
    u = current_user()
    return jsonify(username=u["username"], role=u["role"], can_edit=can_edit(u))

@app.route("/api/data")
@login_required
def api_data():
    return jsonify(data=read_workbook(), version=file_version(), colorder=load_colorder())

@app.route("/api/version")
@login_required
def api_version():
    return jsonify(version=file_version())

@app.route("/api/update", methods=["POST"])
@login_required
def api_update():
    u = current_user()
    if not can_edit(u):
        return jsonify(ok=False, error="You do not have edit permission."), 403
    body = request.get_json(force=True)
    sheet = body.get("sheet"); month = body.get("month"); values = body.get("values", {})
    if not sheet or not month:
        return jsonify(ok=False, error="Missing sheet or month"), 400
    try:
        clean = {}
        for k, v in values.items():
            if v in ("", None):
                clean[k] = None
            else:
                clean[k] = float(v)
        write_month_row(sheet, month, clean)
        return jsonify(ok=True, version=file_version())
    except PermissionError:
        return jsonify(ok=False, error="Could not save — the Excel file is open. Close it in Excel and try again."), 423
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 400

@app.route("/api/colorder", methods=["POST"])
@admin_required
def api_colorder_save():
    b = request.get_json(force=True)
    sheet = b.get("sheet")
    order = b.get("order")
    if not sheet or not isinstance(order, list):
        return jsonify(ok=False, error="Bad request"), 400
    data = read_workbook()
    if sheet not in data:
        return jsonify(ok=False, error="Unknown sheet"), 404
    valid = [k for k in data[sheet][0].keys() if k != "Month"]
    order = [c for c in order if c in valid]
    with _cfg_lock:
        d = load_colorder()
        d[sheet] = order
        save_colorder(d)
    return jsonify(ok=True)

# ---- User management API (admin only) ----
@app.route("/api/users", methods=["GET"])
@admin_required
def api_users_list():
    rows = sorted(_read_users(), key=lambda x: x["id"])
    return jsonify(users=[{"id": r["id"], "username": r["username"], "role": r["role"],
                           "can_edit": int(r["can_edit"]), "created": r["created"]} for r in rows])

@app.route("/api/users", methods=["POST"])
@admin_required
def api_users_create():
    b = request.get_json(force=True)
    un = (b.get("username") or "").strip()
    pw = b.get("password") or ""
    role = "admin" if b.get("role") == "admin" else "user"
    ce = 1 if (b.get("can_edit") or role == "admin") else 0
    if not un or not pw:
        return jsonify(ok=False, error="Username and password required"), 400
    with _users_lock:
        rows = _read_users()
        if any(str(r["username"]).lower() == un.lower() for r in rows):
            return jsonify(ok=False, error="Username already exists"), 400
        rows.append({"id": _next_id(rows), "username": un,
                     "pw_hash": generate_password_hash(pw), "role": role, "can_edit": ce,
                     "created": datetime.now().isoformat(timespec="seconds")})
        try:
            _write_users(rows)
        except PermissionError:
            return jsonify(ok=False, error="Could not save — users_config.xlsx is open. Close it and retry."), 423
    return jsonify(ok=True)

@app.route("/api/users/<int:uid>", methods=["POST"])
@admin_required
def api_users_modify(uid):
    b = request.get_json(force=True)
    action = b.get("action")
    me = current_user()
    with _users_lock:
        rows = _read_users()
        target = next((r for r in rows if r["id"] == uid), None)
        if not target:
            return jsonify(ok=False, error="User not found"), 404
        if action == "delete":
            if target["id"] == me["id"]:
                return jsonify(ok=False, error="You cannot delete yourself"), 400
            if target["role"] == "admin" and sum(1 for r in rows if r["role"] == "admin") <= 1:
                return jsonify(ok=False, error="Cannot delete the last admin"), 400
            rows = [r for r in rows if r["id"] != uid]
        elif action == "toggle_edit":
            target["can_edit"] = 0 if target["can_edit"] else 1
        elif action == "set_role":
            role = "admin" if b.get("role") == "admin" else "user"
            target["role"] = role
            if role == "admin":
                target["can_edit"] = 1
        elif action == "reset_pw":
            pw = b.get("password") or ""
            if not pw:
                return jsonify(ok=False, error="Password required"), 400
            target["pw_hash"] = generate_password_hash(pw)
        else:
            return jsonify(ok=False, error="Unknown action"), 400
        try:
            _write_users(rows)
        except PermissionError:
            return jsonify(ok=False, error="Could not save — users_config.xlsx is open. Close it and retry."), 423
    return jsonify(ok=True)

# ---- Excel export (filters applied) ----
def _send_xlsx(df_map, fname):
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        for name, df in df_map.items():
            df.to_excel(xw, sheet_name=name[:31], index=False)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export/table")
@login_required
def export_table():
    sheet = request.args.get("sheet")
    year = request.args.get("year")
    month = request.args.get("month")  # optional "MM"
    data = read_workbook()
    if sheet not in data:
        return abort(404)
    df = pd.DataFrame(data[sheet])
    if year:
        df = df[df["Month"].astype(str).str.startswith(str(year))]
    if month:
        df = df[df["Month"].astype(str).str.slice(5, 7) == str(month)]
    safe = sheet.replace(" ", "_")
    tag = (year or "all") + ("-" + month if month else "")
    return _send_xlsx({sheet: df}, f"{safe}_{tag}.xlsx")

@app.route("/api/export/chart")
@login_required
def export_chart():
    sheet = request.args.get("sheet")
    month = request.args.get("month")
    year = request.args.get("year")
    data = read_workbook()
    if sheet not in data:
        return abort(404)

    all_months = (str(month) == "all")
    if all_months:
        yr = str(year or "")
        rows = [r for r in data[sheet] if str(r["Month"]).startswith(yr)]
        if not rows:
            return abort(404)
        month_keys = [r["Month"] for r in rows]
        period_label = f"All months {yr}"
        month_str, year_str = "All", yr
        valcols = [datetime(int(mk.split("-")[0]), int(mk.split("-")[1]), 1).strftime("%b") for mk in month_keys]
    else:
        row = next((r for r in data[sheet] if r["Month"] == month), None)
        if not row:
            return abort(404)
        month_keys = [month]
        y, m = month.split("-")
        period_label = datetime(int(y), int(m), 1).strftime("%b %Y")
        month_str = datetime(int(y), int(m), 1).strftime("%B")
        year_str = y
        valcols = [f"kWh ({period_label})"]

    rowmap = {r["Month"]: r for r in data[sheet]}
    is_solar = bool(re.match(r"^\s*solar\s*$", str(sheet), re.I))

    if is_solar:
        AREA_MAP = {"45Mtr": "45Mtr", "D12 P-I": "D12 Phase-I", "D12 P-II": "D12 Phase-II",
                    "D12 P-III": "D12 Phase-III", "Phase-4A": "Phase-4",
                    "Yarnshed": "Yarnshed", "Solar-Park": "SolarPark"}
        all_cols = [c for c in data[sheet][0].keys() if c != "Month"]
        inv_cols = [c for c in all_cols if re.match(r"^\s*Inv\b", str(c))]
        headers = ["Area", "Name"] + valcols
        table = []
        for c in inv_cols:
            mt = re.search(r"\(([^)]+)\)", str(c))
            area = AREA_MAP.get(mt.group(1).strip(), mt.group(1).strip()) if mt else ""
            table.append([area, c] + [rowmap.get(mk, {}).get(c) for mk in month_keys])
    else:
        order = load_colorder().get(sheet) or [k for k in data[sheet][0].keys() if k != "Month"]
        cols = [c for c in order if c != "Month" and c in data[sheet][0]]
        cols += [c for c in data[sheet][0] if c != "Month" and c not in cols]
        headers = ["Name"] + valcols
        table = [[c] + [rowmap.get(mk, {}).get(c) for mk in month_keys] for c in cols]

    # ---- write workbook with a title block ----
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws["A1"] = sheet
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Month: {month_str}"
    ws["A2"].font = Font(bold=True)
    ws["C2"] = f"Year: {year_str}"
    ws["C2"].font = Font(bold=True)
    hdr = 4
    for j, h in enumerate(headers, 1):
        ws.cell(row=hdr, column=j, value=h).font = Font(bold=True)
    for i, rv in enumerate(table, hdr + 1):
        for j, v in enumerate(rv, 1):
            ws.cell(row=i, column=j, value=v)
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 32)

    fname = re.sub(r'[\\/:*?"<>|]', "-", f"{sheet} {period_label}") + ".xlsx"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export/grid", methods=["POST"])
@login_required
def export_grid():
    b = request.get_json(force=True)
    title = str(b.get("title") or "Export")
    month_str = str(b.get("month") or "")
    year_str = str(b.get("year") or "")
    headers = b.get("headers") or []
    rows = b.get("rows") or []
    sheetname = re.sub(r'[\\/:*?\[\]]', "-", str(b.get("sheetname") or "Data"))[:31] or "Data"
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Month: {month_str}"
    ws["A2"].font = Font(bold=True)
    ws["C2"] = f"Year: {year_str}"
    ws["C2"].font = Font(bold=True)
    hdr = 4
    for j, h in enumerate(headers, 1):
        ws.cell(row=hdr, column=j, value=h).font = Font(bold=True)
    for i, rv in enumerate(rows, hdr + 1):
        for j, v in enumerate(rv, 1):
            ws.cell(row=i, column=j, value=v)
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 32)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"{title}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    init_users()
    app.run(host="0.0.0.0", port=5010, debug=False)
else:
    # ensure the users workbook/admin exist when served via a WSGI server
    init_users()
